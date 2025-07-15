import os
import pandas as pd
import openai
import shutil
from openpyxl import load_workbook
import tqdm

# === 配置 ===
original_file = "/home/gzy/rag-biomap/dataset/VTE-PTE-CTEPH研究数据库.xlsx"
new_file_dir = "/home/gzy/rag-biomap/dataset/test_two"
# new_file_path = os.path.join(new_file_dir, "VTE-PTE-CTEPH研究数据库-实验室检查1.xlsx")
# standard_data_path = os.path.join(new_file_dir, "标准数据.xlsx")
new_file_path = os.path.join(new_file_dir, "VTE-PTE-CTEPH研究数据库-实验室检查2.xlsx")
standard_data_path = os.path.join(new_file_dir, "标准数据2.xlsx")
non_standard_file = "/home/gzy/rag-biomap/dataset/实验室检查数据.1.xlsx"
result_dir = "/home/gzy/rag-biomap/dataset/Results"

# OpenAI client 配置
client = openai.OpenAI(
    base_url="http://172.16.55.171:7010/v1",
    api_key="sk-cairi"
)

prompt_template = r"""
你是一个专业的医疗领域数据对齐助手，擅长将非标准化的实验室检查指标名称映射到标准化的定义。

你的任务是接收一个查询检验指标名（query）和一个候选标准指标名列表（candidates），然后从candidates中找到与query最匹配的指标名及其对应的置信度分数。

请注意以下匹配规则：
1. **完全匹配优先**：如果query与candidates中的某个指标名完全相同（包括大小写），直接返回1.0分
2. **忽略非关键修饰词**：可以忽略（query）中的符号信息，比如★以及*等，但需保留核心指标名称
3. **同义词匹配**：识别医学术语中的同义词关系（如"血红蛋白"和"血色素"）
4. **缩写匹配**：能够处理常见医学缩写（如"WBC"匹配"白细胞计数"）
5. **置信度分数**：分数范围为(0, 1.0]，1.0表示完美匹配
6. **无匹配处理**：若无合理匹配则返回N/A和0.0

返回格式必须严格为JSON，包含 `matched_field_name` 和 `score` 两个字段。不要包含```json和任何额外的解释、说明或报错信息。


---

**输入示例**：
Query: "血红蛋白(*HGB)"
Candidates: ["血红蛋白（Hb）", "血清白蛋白（ALB）", "肌钙蛋白I（cTNI）"]

**输出示例**：
{
  "matched_field_name": "血红蛋白（Hb）",
  "score": 0.9
}

---

## 当前任务

* Query: {{query}}
* Candidates: {{candidates}}

请返回JSON格式的匹配结果：
"""
# === 步骤 1: 复制原始文件并添加录入1列 ===
shutil.copyfile(original_file, new_file_path)
wb = load_workbook(new_file_path)
ws = wb["实验室检查"]

header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
name_col_idx = header_row.index("名称") + 1
insert_col_idx = name_col_idx + 1
ws.insert_cols(insert_col_idx)
ws.cell(row=1, column=insert_col_idx, value="录入1")
wb.save(new_file_path)

# === 步骤 2: 提取标准字段 ===
df_standard = pd.read_excel(original_file, sheet_name="实验室检查", usecols="E")
df_standard.to_excel(standard_data_path, index=False)

# === 步骤 3: 提取指定唯一号下的非标准术语 ===
df_non_standard_all = pd.read_excel(non_standard_file)
df_filtered = df_non_standard_all[df_non_standard_all["唯一号"] == "zy@ZY050002401765"]
query_terms = df_filtered[["检验指标名", "检验指标值"]].dropna().values.tolist()

# === 步骤 4: 调用 LLM 匹配 ===
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
from tqdm import tqdm
import json

def llm_match_single(query, value, candidates):
    filled_prompt = prompt_template.replace("{{query}}", query).replace("{{candidates}}", str(candidates))
    try:
        response = client.chat.completions.create(
            model="CAIRI-LLM-reasoner",
            messages=[{"role": "user", "content": filled_prompt}],
            temperature=0,
            presence_penalty=1.5,
            extra_body={"min_p": 0},
        )
        result = json.loads(response.choices[0].message.content.strip())
        matched = result.get("matched_field_name", "N/A")
        return matched, value
    except Exception:
        return "错误", "错误"
candidates = df_standard["名称"].dropna().astype(str).tolist()
# === 多线程执行 LLM 匹配（替换原 for 循环） ===
match_map = defaultdict(list)
with ThreadPoolExecutor(max_workers=10) as executor:
    futures = {
        executor.submit(llm_match_single, query, value, candidates): (query, value)
        for query, value in query_terms
    }
    for future in tqdm(as_completed(futures), total=len(futures), desc="🧠 LLM多线程匹配中", ncols=80):
        matched, value = future.result()
        match_map[matched].append(value)
# import json
# from collections import defaultdict
# from tqdm import tqdm
#
# candidates = df_standard["名称"].dropna().astype(str).tolist()
# match_map = defaultdict(list)
#
# for query, value in tqdm(query_terms, desc="LLM匹配中", ncols=80):
#     filled_prompt = prompt_template.replace("{{query}}", query).replace("{{candidates}}", str(candidates))
#     try:
#         response = client.chat.completions.create(
#             model="CAIRI-LLM-reasoner",
#             messages=[{"role": "user", "content": filled_prompt}],
#             temperature=0,
#             presence_penalty=1.5,
#             extra_body={"min_p": 0},
#         )
#         result = json.loads(response.choices[0].message.content.strip())
#         matched = result.get("matched_field_name", "N/A")
#         if matched != "N/A":
#             match_map[matched].append(value)
#         else:
#             match_map[matched].append("N/A")
#     except Exception as e:
#         match_map["错误"].append("错误")


# === 步骤 5: 写入到“录入1”列 ===
df_target = pd.read_excel(new_file_path, sheet_name="实验室检查")
if "录入1" not in df_target.columns:
    df_target.insert(name_col_idx, "录入1", "")

for idx, row in df_target.iterrows():
    field_name = str(row["名称"]).strip()
    if field_name in match_map:
        values = match_map[field_name]
        df_target.at[idx, "录入1"] = "--".join(map(str, values))
    elif field_name == "nan":
        df_target.at[idx, "录入1"] = "NA"
    elif any(field_name == m for m in match_map.keys() if m != "错误" and m != "N/A"):
        continue
    elif field_name not in match_map:
        df_target.at[idx, "录入1"] = "NA"
    else:
        df_target.at[idx, "录入1"] = "错误"

with pd.ExcelWriter(new_file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df_target.to_excel(writer, sheet_name="实验室检查", index=False)

print("✅ 所有步骤已完成。")

